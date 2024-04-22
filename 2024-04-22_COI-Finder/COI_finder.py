import os
import openpyxl
import pandas as pd

def find_matching_words_in_xlsx(directory, matching_words):
    # List to store rows containing matching words and their corresponding file names
    matching_rows = []

    # Iterate through each file in the directory
    for filename in os.listdir(directory):
        if filename.endswith('.xlsx'):
            filepath = os.path.join(directory, filename)

            # Open the Excel file
            workbook = openpyxl.load_workbook(filepath)

            # Iterate through each sheet in the workbook
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]

                # Iterate through each row in the sheet
                for row in sheet.iter_rows(values_only=True):
                    for cell in row:
                        if cell and any(word.lower() in str(cell).lower() for word in matching_words):
                            # If any matching word is found in the cell, store the row and filename
                            matching_rows.append((filename, row))
                            break  # Move to the next row

    return matching_rows

def save_to_excel(output, output_file):
    # Create a DataFrame from the output
    df = pd.DataFrame(output, columns=["File Name", "Matching Row"])

    # Sort the DataFrame by the 'File Name' column in ascending order
    df.sort_values(by='File Name', ascending=True, inplace=True)

    # Save the DataFrame to an Excel file
    df.to_excel(output_file, index=False)

# Example usage
directory = "sheets"
matching_words = ["COI", "Conflict of Interest"]

result = find_matching_words_in_xlsx(directory, matching_words)
output_file = "COI_documents.xlsx"
save_to_excel(result, output_file)
print(f"Output saved to '{output_file}'")